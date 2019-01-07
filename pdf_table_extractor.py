#!/usr/bin/python3

import camelot
import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
import codecs
import os
import sys
import xlsxwriter

'''
command to split pdf : pdftk <input.pdf> burst
https://camelot-py.readthedocs.io/en/master/
pip3 install camelot-py[cv]
'''

if len(sys.argv) == 1:

    print ("")
    print ("")
    print ("PDF filename not passed as PDF. Please pass PDF filename as first argument OR Parameter")
    print ("")
    print ("")
    sys.exit()


pdf = sys.argv[1]

os.system('mkdir input')
os.system('cp ' + pdf + ' input/')
os.system('chmod -R 777 input')
#Split pdf using pdftk cmd
os.system('cd input && pdftk ' + pdf + ' burst')
os.remove(os.path.join('input',pdf))
os.system('rm -Rf input/*.txt')
os.system('chmod -R 777 input')

#Converting the splitted PDF into list
pdflist = sorted(os.listdir('input'))

#Create Excel Workbook
workbook = xlsxwriter.Workbook('tables.xlsx')
worksheet = workbook.add_worksheet()
workbook.close()
os.system('chmod -R 777 tables.xlsx')

#Load excel Workbook using openpyxl
book = load_workbook('tables.xlsx')
writer = ExcelWriter('tables.xlsx', engine='openpyxl') 	
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

#Iterating through the PDF list
for i in range(0,len(pdflist)):

    pdfn = pdflist[i].replace("pg_","").replace(".pdf","")

    #extracting table from PDF
    tables = camelot.read_pdf(os.path.join('input',pdflist[i]))

    #check if pdf contains table to be extracted
    if "1" in str(tables):
        print ("table found in " + pdflist[i])
        #converting a pdf into a DataFrame
        tables[0].df.to_excel(writer,sheet_name="table_" + str(pdfn),index=False,header=False)
	
writer.save()
